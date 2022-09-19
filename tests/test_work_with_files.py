import pytest
import zipfile
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import xlrd


def create_Zip():
    with zipfile.ZipFile('../resources/resources.zip', 'w') as myzip:
        myzip.write('../resources/items.xls')
        myzip.write('../resources/items2.xlsx')
        myzip.write('../resources/book.csv')
        myzip.write('../resources/ISTQB_CTFL_PT_Syllabus_2018_GA.pdf')
        myzip.write('../resources/readme33.txt')


def extract_Zip():
    with zipfile.ZipFile('../resources/resources.zip') as zip_:
        print(zip_.namelist())
        zip_.extractall('../resources/archive')


create_Zip()
extract_Zip()


def test_matches_value_after_extract_in_txt():
    with open('../resources/archive/resources/readme33.txt', 'r+') as txtfile:
        data = txtfile.read()
        print(data)
        assert '1111111111\n22222\n33333\n' == data

def test_read_value_txt_in_archive():
    with zipfile.ZipFile('../resources/resources.zip') as myzip:
        with open('../resources/archive/resources/readme33.txt', 'r+') as txtfile:
            data = txtfile.read()
            print(data)
            assert '1111111111\n22222\n33333\n' == data

def test_matches_value_after_extract_in_csv():
    with open('../resources/archive/resources/book.csv', 'r+') as csvfile:
        data = ''
        table = csv.reader(csvfile)
        for line_no, line in enumerate(table, 1):
            if line_no == 2:
                data = str(line)
        assert "['100', 'Aigneis', 'Connelly', 'Aigneis.Connelly@yopmail.com', 'Aigneis.Connelly@gmail.com', 'firefighter']" == data

def test_read_value_csv_in_archive():
    with zipfile.ZipFile('../resources/resources.zip') as myzip:
        with open('../resources/archive/resources/book.csv', 'r+') as csvfile:
            table = csv.reader(csvfile)
            for line_no, line in enumerate(table, 1):
                if line_no == 2:
                    data = str(line)
            assert "['100', 'Aigneis', 'Connelly', 'Aigneis.Connelly@yopmail.com', 'Aigneis.Connelly@gmail.com', 'firefighter']" == data

def test_matches_value_after_extract_in_pdf():
    pdf_docs = PdfReader('../resources/archive/resources/ISTQB_CTFL_PT_Syllabus_2018_GA.pdf')
    get_text = pdf_docs.pages[1].extract_text()
    assert 'Version 2018 Page 2 of 59 9 December 2018 ' in get_text

def test_read_value_pdf_in_archive():
    with zipfile.ZipFile('../resources/resources.zip') as myzip:
        pdf_docs = PdfReader('../resources/ISTQB_CTFL_PT_Syllabus_2018_GA.pdf')
        get_text = pdf_docs.pages[1].extract_text()
        assert 'Version 2018 Page 2 of 59 9 December 2018 ' in get_text

def test_matches_value_after_extract_in_xls():
    xls_table = xlrd.open_workbook('../resources/archive/resources/items.xls')
    sheet1 = xls_table.sheet_by_index(0)
    assert 354563.0 == sheet1.cell_value(rowx=2, colx=1)

def test_read_value_xls_in_archive():
    with zipfile.ZipFile('../resources/resources.zip') as myzip:
        xls_table = xlrd.open_workbook('../resources/items.xls')
        sheet1 = xls_table.sheet_by_index(0)
        assert 354563.0 == sheet1.cell_value(rowx=2, colx=1)

def test_matches_value_after_extract_in_xlsx():
    book = load_workbook('../resources/archive/resources/items2.xlsx')
    sheets = book.active
    assert 'Код товара' == sheets.cell(row=1, column=1).value

def test_read_value_xlsx_in_archive():
    with zipfile.ZipFile('../resources/resources.zip') as myzip:
        book = load_workbook('../resources/items2.xlsx')
        sheets = book.active
        assert 'Код товара' == sheets.cell(row=1, column=1).value
