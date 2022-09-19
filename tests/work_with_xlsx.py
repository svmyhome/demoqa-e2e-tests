from openpyxl import load_workbook

book = load_workbook('../resources/items2.xlsx')
sheets = book.active
print(sheets)
print(sheets.cell(row=1, column=1).value)
for x in range(1, 10):
    for y in range(1, 10):
        print(sheets.cell(row=x, column=y).value)
